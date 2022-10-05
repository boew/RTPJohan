import pandas as pd
import datetime as dt
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
class MhdTracker:
    """
    """
    def __init__(self, x, iMax, nPulsesIn, pathP='../web/'):
        self.plot = None
        self.x = x
        self.tmpPath = Path(f'{pathP}tmp_{x}plot.svg')
        self.realPath = Path(f'{pathP}{x}plot.svg')
        self.txtPath = Path(f'{pathP}{x}log.txt')
        self.xlsxPath = Path(f'{pathP}{x}log.xlsx')
        if (self.xlsxPath.exists()):
            newName = "old_" + dt.datetime.now().strftime("%Y-%m-%d_%H%M%S") + self.xlsxPath.name
            self.xlsxPath.rename(self.xlsxPath.with_name(newName))
        self.xlsxWb = Workbook()     
        self.xlsxWb.save(self.xlsxPath)
        self.iMax=iMax
        self.i=0
        self.acc_ns=0
        self.acc_cnt=0
        self.nPulsesIn = nPulsesIn
        self.series=pd.Series([0] * self.iMax)
        self.t01_xlsxPath = Path(f'{pathP}{x}t01.xlsx')
        if (self.t01_xlsxPath.exists()):
            newName = "old_" + dt.datetime.now().strftime("%Y-%m-%d_%H%M%S") + self.t01_xlsxPath.name
            self.t01_xlsxPath.rename(self.t01_xlsxPath.with_name(newName))
        self.t01_xlsxWb = Workbook()     
        self.t01_xlsxWb.save(self.t01_xlsxPath)
        self.formatString = "%Y-%m-%d %H:%M:%S"
        # 2022-09-11 12:12:01
        #    Y  m  d  H  M  S
        if (not MhdTracker.mhd_t0):
            MhdTracker.mhd_t0 = dt.datetime.now()

    def ns2power(self, nano_s):
        return 3.6E9 * self.nPulsesIn / nano_s

    def ns2time(self, nano_s):
        return (MhdTracker.mhd_t0 + dt.timedelta(milliseconds = (self.acc_ns + nano_s) // 1e6)).strftime("%H:%M")

    def ns2datetime(self, nano_s):
        t = MhdTracker.mhd_t0 + dt.timedelta(milliseconds = (self.acc_ns + nano_s) / 1e6)
        datetime = t.strftime(self.formatString)
        year = t.strftime("%Y")
        month = t.strftime("%m")
        day = t.strftime("%d")
        hour = t.strftime("%H")
        minute = t.strftime("%M")
        return [datetime, year, month, day, hour, minute]
       
    def update(self, tns1, tns0=0):
        self.t01_xlsxWb.active.append([tns0, tns1])
        self.t01_xlsxWb.save(self.t01_xlsxPath)
        tnsDiff = tns1 - tns0
        if (0 == self.i):
            self.acc_ns += self.series.sum()
            self.acc_cnt += 1
        self.series[self.i] = tnsDiff
        self.i += 1
        self.i %= self.iMax
        return (0 == self.i)

    def doLogTxt(self):
        with open(self.txtPath, 'a') as logfile:
            print(self.series, file=logfile)

    def doLogXlsx(self):
        local_acc_ns = 0
        for tns in self.series:
            local_acc_ns += tns
            local_t = self.ns2datetime(local_acc_ns)
            row = local_t
            row.insert(0,tns)
            row.insert(1,1)
            row.append(dt.datetime.now().strftime(self.formatString))
            self.xlsxWb.active.append(row)
        self.xlsxWb.save(self.xlsxPath)
            
    def doPlot(self, toFile=False):
        if (self.plot):
            plt.close(self.plot.get_figure())
        self.sdf = pd.DataFrame({'time': list(map(self.ns2time, self.series.cumsum())),
                                 'power': list(map(self.ns2power, self.series))})
        pdt0 = None
        for pdidx in self.sdf.index:
            pdt = self.sdf.at[pdidx,'time']
            if pdt0 == pdt:
                self.sdf.at[pdidx,'time'] = ""
            else:
                pdt0 = pdt

        self.plot = self.sdf.plot('time','power', kind='bar')
        self.tTime = MhdTracker.mhd_t0 + dt.timedelta(milliseconds = self.acc_ns // 1e6)
        self.pTitle  = f'{self.x}plot {self.acc_cnt}@'
        self.pTitle += self.tTime.strftime(self.formatString)
        print(self.pTitle)
        self.plot.set_title(self.pTitle)
        if (toFile):
            plt.savefig(self.tmpPath)
            plt.cla()
            self.tmpPath.replace(self.realPath)
        else:
            plt.show()

if '__main__' == __name__ :
    mT=MhdTracker('m', 16, pathP='../test/')       
    for i in range(16):
        j=i+3
        print(mT.update(j*3-j))
    print(mT.series)
    mT.doPlot()
    mT.doLogXlsx()
    mT.doLogTxt()
