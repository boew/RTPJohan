import sys
import pandas as pd
import datetime as dt
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import struct

class MhdTracker:
    """
    """
    mhd_t0 = None
    def __init__(self, x, iMax, nPulsesIn, pathP='../web/'):
        self.plot = None
        self.x = x
        self.tmpPath = Path(f'{pathP}tmp_{x}plot.svg')
        self.realPath = Path(f'{pathP}{x}plot.svg')
        self.txtPath = Path(f'{pathP}{x}log.txt')
        self.xlsxPath = Path(f'{pathP}{x}log.xlsx')
        if (self.xlsxPath.exists()):
            wb = load_workbook(self.xlsxPath)
        else:
            wb = Workbook()     
        self.xlsxWb = wb
        self.xlsxWb.save(self.xlsxPath)
        self.iMax=iMax
        self.i=0
        self.acc_ns=0
        self.acc_cnt=0
        self.nPulsesIn = nPulsesIn
        self.series=pd.Series([0] * self.iMax)
        self.formatString = "%Y-%m-%d %H:%M:%S"
        # 2022-09-11 12:12:01
        #    Y  m  d  H  M  S

    def ns2power(self, nano_s):
        return 3.6E9 * self.nPulsesIn / nano_s

    def ns2time(self, nano_s):
        return (MhdTracker.mhd_t0 + dt.timedelta(milliseconds = (self.acc_ns + nano_s) // 1e6)).strftime("%H:%M")

    def ns2datetime(self, nano_s):
        t = MhdTracker.mhd_t0 + dt.timedelta(milliseconds = (self.acc_ns + nano_s) / 1e6)
        datetime = t.strftime(self.formatString)
        year = t.strftime("%Y")
        month = t.strftime("%m")
        day = t.strftime("%d")
        hour = t.strftime("%H")
        minute = t.strftime("%M")
        return [datetime, year, month, day, hour, minute]
    
    def update(self, tnsDiff):
        self.series[self.i] = tnsDiff
        self.i += 1
        self.i %= self.iMax
        if (0 == self.i):          
            if (not MhdTracker.mhd_t0):
                MhdTracker.mhd_t0 = dt.datetime.now()
            self.acc_ns += self.series.sum()
            self.acc_cnt += 1
        return (0 == self.i)    

    def doLogTxt(self):
        with open(self.txtPath, 'a') as logfile:
            print(self.series, file=logfile)

    def doLogXlsx(self):
        local_acc_ns = 0
        for tns in self.series:
            local_acc_ns += tns
            local_t = self.ns2datetime(local_acc_ns)
            row = local_t
            row.insert(0,tns)
            row.insert(1,1)
            self.xlsxWb.active.append(row)
        self.xlsxWb.save(self.xlsxPath)
            
    def doPlot(self, toFile=False):
        if (self.plot):
            plt.close(self.plot.get_figure())
        self.sdf = pd.DataFrame({'time': list(map(self.ns2time, self.series.cumsum())),
                                 'power': list(map(self.ns2power, self.series))})
        pdt0 = None
        for pdidx in self.sdf.index:
            pdt = self.sdf.at[pdidx,'time']
            if pdt0 == pdt:
                self.sdf.at[pdidx,'time'] = ""
            else:
                pdt0 = pdt

        self.plot = self.sdf.plot('time','power', kind='bar')
        self.tTime = MhdTracker.mhd_t0 + dt.timedelta(milliseconds = self.acc_ns // 1e6)
        self.pTitle  = f'{self.x}plot {self.acc_cnt}@'
        self.pTitle += self.tTime.strftime(self.formatString)
        print(self.pTitle)
        self.plot.set_title(self.pTitle)
        if (toFile):
            plt.savefig(self.tmpPath)
            plt.cla()
            self.tmpPath.replace(self.realPath)
        else:
            plt.show()

if '__main__' == __name__ :
    mT = MhdTracker('m', 16,  1, './')      # 16 * ~5s <-> ~80 s  
    hT = MhdTracker('h', 45, 16, './')      # 45 * ~80 s <-> ~1 h 
    dT = MhdTracker('d', 24, 45, './')      # 24 * ~1 h <-> ~1 day
    formatString = "%Y-%m-%d; %H:%M:%S"
    def processBlink(tns0, tns1):
        if mT.update(tns1-tns0):
            mT.doPlot(toFile=True)
            mT.doLogTxt()
            mT.doLogXlsx()
            if hT.update(mT.series.sum()):
                hT.doPlot(True)
                if dT.update(hT.series.sum()):
                    dT.doPlot(True)

    print(f'{sys.argv[0]} @ {dt.datetime.now().strftime(formatString)}')
    with open('mktest4.txt','rb') as binaryFile:
        tt = binaryFile.read(16)
        while 16 == len(tt):
            (tns0,tns1) = struct.unpack('QQ',tt)
            #print(tns0,tns1)
            processBlink(tns0,tns1)
            tt = binaryFile.read(16)
